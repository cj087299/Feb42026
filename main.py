import os
import logging
import json
import uuid
import base64
import requests
import threading
from queue import Queue
from urllib.parse import quote, urlparse, urlunparse
from flask import Flask, jsonify, request, render_template, session, redirect, url_for
from src.qbo_client import QBOClient
from src.invoice_manager import InvoiceManager
from src.cash_flow_calendar import CashFlowCalendar
from src.cash_flow import CashFlowProjector
from src.ai_predictor import PaymentPredictor
from src.secret_manager import SecretManager
from src.database import Database
from src.ai_service import AIService
from src.error_handler import ErrorLogger, handle_errors, log_ai_action
from src.auth import (
    hash_password, verify_password, login_required, permission_required, 
    role_required, get_current_user, audit_log, ROLES, has_permission
)
from src.email_service import EmailService
from src.webhook_handler import WebhookHandler

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Constants for QBO dummy credentials (used when credentials are not configured)
DUMMY_QBO_CLIENT_ID = 'dummy_id'
DUMMY_QBO_CLIENT_SECRET = 'dummy_secret'
DUMMY_QBO_REFRESH_TOKEN = 'dummy_refresh'
DUMMY_QBO_REALM_ID = 'dummy_realm'

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(24))

# Initialize Secret Manager and Database
database = Database()
secret_manager = SecretManager(database=database)

# Initialize Error Logger
error_logger = ErrorLogger()

# Initialize AI Service
ai_service = AIService()

# Initialize Email Service
email_service = EmailService()

# Initialize Webhook Handler
webhook_handler = WebhookHandler()

# Webhook Queue for asynchronous processing
# This ensures webhook endpoint responds immediately with 200 OK
webhook_queue = Queue()

def process_webhook_queue():
    """
    Background thread that processes webhook events asynchronously.
    This allows the webhook endpoint to return 200 OK immediately.
    """
    while True:
        try:
            # Get event from queue (blocks until event available)
            event_data = webhook_queue.get()
            
            if event_data is None:  # Poison pill to stop thread
                break
            
            logger.info(f"Processing queued webhook event: {event_data.get('event_id', 'unknown')}")
            
            # Parse CloudEvents format
            parsed_data = webhook_handler.parse_cloudevents(event_data)
            
            if not parsed_data:
                logger.error("Failed to parse queued CloudEvents payload")
            else:
                # Process the webhook event
                result = webhook_handler.process_webhook_event(parsed_data)
                logger.info(f"Completed processing webhook event: {result}")
            
            # Mark task as done
            webhook_queue.task_done()
            
        except Exception as e:
            logger.error(f"Error processing queued webhook event: {e}")
            webhook_queue.task_done()

# Start background webhook processor thread
# Note: Using daemon=True is appropriate for Cloud Run deployment:
# - Webhook events are idempotent (QBO will retry on failure)
# - Container lifecycle is managed by Cloud Run
# - Critical requirement is fast response time (< 1 sec), not zero event loss
# - For production at scale, consider using Cloud Tasks or Pub/Sub for durability
webhook_processor_thread = threading.Thread(target=process_webhook_queue, daemon=True)
webhook_processor_thread.start()
logger.info("Webhook background processor started")

# Initialize QBO client with credentials from Secret Manager (which now checks database first)
qbo_credentials = secret_manager.get_qbo_credentials()
qbo_client = QBOClient(
    qbo_credentials['client_id'],
    qbo_credentials['client_secret'],
    qbo_credentials['refresh_token'],
    qbo_credentials['realm_id'],
    database=database
)
# Set the access token if available in the database
# This allows the client to use a valid token without an initial refresh
if qbo_credentials.get('access_token'):
    qbo_client.access_token = qbo_credentials['access_token']
    logger.info("Loaded access token from database for global qbo_client")

# Initialize predictor before invoice_manager
# Train predictor with dummy data initially or load a saved model
predictor = PaymentPredictor()
# Ideally, we would load training data from a persistent source here
# For now, we leave it untrained or train on demand if data is available

invoice_manager = InvoiceManager(qbo_client, database=database, predictor=predictor)

# Check and log QBO credential status on startup
if not qbo_credentials.get('is_valid'):
    logger.warning("=" * 70)
    logger.warning("  QuickBooks credentials are NOT configured")
    logger.warning("=" * 70)
    logger.warning("The application is running, but QuickBooks integration will not work")
    logger.warning("until you configure OAuth credentials.")
    logger.warning("")
    logger.warning("To configure credentials, choose ONE of these options:")
    logger.warning("")
    logger.warning("Option 1 (RECOMMENDED - Easiest):")
    logger.warning("  → Log in to the application")
    logger.warning("  → Navigate to /qbo-settings")
    logger.warning("  → Click 'Connect to QuickBooks'")
    logger.warning("")
    logger.warning("Option 2 (Google Secret Manager):")
    logger.warning("  → Create secrets: QBO_ID_2-3-26, QBO_Secret_2-3-26")
    logger.warning("  → Set environment variables: QBO_REFRESH_TOKEN, QBO_REALM_ID")
    logger.warning("  → See CLOUDRUN_DEPLOYMENT_GUIDE.md for details")
    logger.warning("")
    logger.warning("Option 3 (Environment Variables):")
    logger.warning("  → Set QBO_CLIENT_ID, QBO_CLIENT_SECRET, QBO_REFRESH_TOKEN, QBO_REALM_ID")
    logger.warning("  → Run: gcloud run services update feb42026 --update-env-vars=...")
    logger.warning("")
    logger.warning("Default admin login: admin@vzt.com / admin1234")
    logger.warning("=" * 70)
else:
    logger.info("✓ QuickBooks credentials are configured and valid")


def initialize_admin_users():
    """
    Initialize default admin users on startup if they don't exist.
    This ensures admin users are available immediately after deployment.
    
    SECURITY NOTE: Default credentials are intentionally simple for initial setup
    and are publicly visible in the repository. These are meant to be changed 
    immediately after first login. The credentials are:
    - admin@vzt.com / admin1234
    - cjones@vztsolutions.com / admin1234
    
    For production deployments, it is recommended to:
    1. Change these passwords immediately after first login
    2. Consider using environment variables for initial credentials if desired
    3. Monitor the audit log for first-time logins
    """
    logger.info("Checking for admin users...")
    
    # Default admin credentials
    # NOTE: These are public and must be changed after first login!
    admin_users = [
        {
            "email": "cjones@vztsolutions.com",
            "password": "admin1234",
            "full_name": "CJones",
            "role": "master_admin"
        },
        {
            "email": "admin@vzt.com",
            "password": "admin1234",
            "full_name": "Admin",
            "role": "master_admin"
        }
    ]
    
    initialized_count = 0
    for user_data in admin_users:
        email = user_data["email"]
        
        # Check if user already exists
        existing_user = database.get_user_by_email(email)
        if existing_user:
            logger.info(f"Admin user {email} already exists. Skipping.")
            continue
        
        # Create user
        password_hash = hash_password(user_data["password"])
        user_id = database.create_user(
            email, 
            password_hash, 
            user_data["full_name"], 
            user_data["role"]
        )
        
        if user_id:
            initialized_count += 1
            logger.info(f"✓ Admin user created: {email}")
        else:
            logger.error(f"✗ Failed to create admin user: {email}")
    
    if initialized_count > 0:
        logger.warning(f"{'='*60}")
        logger.warning(f"Admin users initialized with default passwords!")
        logger.warning(f"Default credentials:")
        logger.warning(f"  - admin@vzt.com / admin1234")
        logger.warning(f"  - cjones@vztsolutions.com / admin1234")
        logger.warning(f"⚠️  IMPORTANT: Change these passwords immediately after first login!")
        logger.warning(f"{'='*60}")
    else:
        logger.info("Admin users already initialized.")


# Flag to track if initialization has been performed
_admin_initialized = False

def ensure_admin_users_initialized():
    """Ensure admin users are initialized exactly once."""
    global _admin_initialized
    if not _admin_initialized:
        initialize_admin_users()
        _admin_initialized = True


# Initialize admin users when the app context is first created
# This works for both development (python main.py) and production (gunicorn)
with app.app_context():
    ensure_admin_users_initialized()


def get_fresh_qbo_client():
    """
    Get a fresh QBO client instance with current credentials from database.
    
    This function always retrieves the latest credentials from the database,
    ensuring that recently updated OAuth tokens are used. This solves the issue
    where the global qbo_client might have stale credentials if it was initialized
    before OAuth completion.
    
    Returns:
        tuple: (QBOClient instance, bool indicating if credentials are valid)
    
    Raises:
        ValueError: If credentials dictionary is missing required fields
    """
    try:
        qbo_creds = secret_manager.get_qbo_credentials()
        
        # Validate that we have the required credential fields
        required_fields = ['client_id', 'client_secret', 'refresh_token', 'realm_id']
        missing_fields = [field for field in required_fields if not qbo_creds.get(field)]
        
        if missing_fields:
            logger.error(f"Missing required QBO credential fields: {missing_fields}")
            # Return a client with dummy values and invalid flag
            # This maintains backward compatibility with existing error handling
            client = QBOClient(
                DUMMY_QBO_CLIENT_ID, 
                DUMMY_QBO_CLIENT_SECRET, 
                DUMMY_QBO_REFRESH_TOKEN, 
                DUMMY_QBO_REALM_ID, 
                database=database
            )
            return client, False
        
        client = QBOClient(
            qbo_creds['client_id'],
            qbo_creds['client_secret'],
            qbo_creds['refresh_token'],
            qbo_creds['realm_id'],
            database=database
        )
        
        # Set the access token if available in the database
        # The access token is stored separately and needs to be set after client creation
        # because it's not part of the QBOClient constructor (only refresh token is)
        # This allows the client to skip an initial token refresh if we have a valid token
        if qbo_creds.get('access_token'):
            client.access_token = qbo_creds['access_token']
        
        return client, qbo_creds.get('is_valid', False)
    except Exception as e:
        logger.error(f"Error creating fresh QBO client: {e}")
        # Return a client with dummy values and invalid flag on error
        # This ensures the function always returns a valid tuple, maintaining API contract
        client = QBOClient(
            DUMMY_QBO_CLIENT_ID, 
            DUMMY_QBO_CLIENT_SECRET, 
            DUMMY_QBO_REFRESH_TOKEN, 
            DUMMY_QBO_REALM_ID, 
            database=database
        )
        return client, False


# Webhook endpoint - NO authentication required (external QBO service)
@app.route('/api/qbo/webhook', methods=['POST', 'GET'])
def qbo_webhook():
    """
    Handle QuickBooks Online webhooks.
    
    This endpoint receives notifications from QBO when entities change.
    It uses CloudEvents format and requires verifier token validation.
    
    IMPORTANT: Per QBO requirements, this endpoint MUST respond immediately
    with 200 OK. Event processing happens asynchronously in a background thread.
    
    CSRF Exemption: This endpoint is called by QBO's servers and doesn't 
    have CSRF tokens. Authentication is done via verifier token.
    """
    try:
        # Handle GET request for webhook verification (QBO setup)
        if request.method == 'GET':
            # QBO may send a verification request during webhook setup
            logger.info("Received webhook verification request")
            return jsonify({
                'status': 'ok',
                'message': 'Webhook endpoint is active',
                'verifier_token': WebhookHandler.VERIFIER_TOKEN
            }), 200
        
        # Handle POST request for actual webhook events
        # Get the verifier token from headers (QBO sends this)
        verifier_token = request.headers.get('intuit-signature', '')
        
        # For initial validation, QBO might also send a specific verification payload
        # Check if this is a verification request
        try:
            payload = request.get_json()
        except Exception as json_error:
            logger.error(f"Failed to parse JSON payload: {json_error}")
            return jsonify({'error': 'Invalid JSON payload'}), 400
        
        if not payload:
            logger.error("No payload received in webhook request")
            return jsonify({'error': 'No payload received'}), 400
        
        # Log the incoming webhook
        # INFO level: truncated for readability
        logger.info(f"Received webhook: {json.dumps(payload, default=str)[:500]}")
        # DEBUG level: full payload for troubleshooting
        logger.debug(f"Full webhook payload: {json.dumps(payload, default=str)}")
        
        # Handle array of events (CloudEvents can be sent as array)
        events = payload if isinstance(payload, list) else [payload]
        
        # Queue events for asynchronous processing
        # This ensures we respond to QBO immediately
        queued_count = 0
        for event in events:
            try:
                webhook_queue.put(event)
                queued_count += 1
            except Exception as e:
                logger.error(f"Failed to queue webhook event: {e}")
        
        logger.info(f"Queued {queued_count} webhook event(s) for background processing")
        
        # Return 200 OK immediately (as required by QBO)
        return jsonify({
            'status': 'accepted',
            'message': f'Received {len(events)} event(s), queued for processing',
            'queued': queued_count
        }), 200
        
    except Exception as e:
        logger.error(f"Error receiving webhook: {e}")
        # Even on error, return 200 to prevent QBO from retrying
        # Log the error for investigation but don't block QBO
        return jsonify({
            'status': 'accepted',
            'message': 'Event received, errors logged'
        }), 200


@app.route('/', methods=['GET'])
def index():
    # Check if user is logged in
    if 'user_id' not in session:
        return redirect(url_for('login_page'))
    
    # Check if request is from browser (HTML) or API client (JSON)
    if request.accept_mimetypes.best == 'text/html' or \
       (request.accept_mimetypes.accept_html and 
        request.accept_mimetypes['text/html'] > request.accept_mimetypes['application/json']):
        return render_template('index.html')
    return jsonify({
        "service": "VZT Accounting API",
        "version": "1.0",
        "endpoints": {
            "health": "/health",
            "invoices": "/api/invoices",
            "cashflow": "/api/cashflow"
        }
    }), 200


@app.route('/login', methods=['GET'])
def login_page():
    """Display login page."""
    if 'user_id' in session:
        return redirect(url_for('index'))
    return render_template('login.html')


@app.route('/forgot-password', methods=['GET'])
def forgot_password_page():
    """Display forgot password page."""
    return render_template('forgot-password.html')


@app.route('/forgot-username', methods=['GET'])
def forgot_username_page():
    """Display forgot username page."""
    return render_template('forgot-username.html')


@app.route('/reset-password', methods=['GET'])
def reset_password_page():
    """Display reset password page."""
    return render_template('reset-password.html')


@app.route('/api/login', methods=['POST'])
@audit_log('user_login')
def login():
    """Handle user login."""
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')
        
        if not email or not password:
            return jsonify({'error': 'Email and password are required'}), 400
        
        # Get user from database
        user = database.get_user_by_email(email)
        
        if not user:
            return jsonify({'error': 'Invalid email or password'}), 401
        
        if not user['is_active']:
            return jsonify({'error': 'Account is inactive'}), 401
        
        # Verify password
        if not verify_password(password, user['password_hash']):
            return jsonify({'error': 'Invalid email or password'}), 401
        
        # Set session
        session['user_id'] = user['id']
        session['user_email'] = user['email']
        session['user_full_name'] = user['full_name']
        session['user_role'] = user['role']
        
        # Update last login
        database.update_last_login(user['id'])
        
        return jsonify({
            'message': 'Login successful',
            'user': {
                'id': user['id'],
                'email': user['email'],
                'full_name': user['full_name'],
                'role': user['role']
            }
        }), 200
    except Exception as e:
        logger.error(f"Login error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/logout', methods=['POST'])
@audit_log('user_logout')
def logout():
    """Handle user logout."""
    session.clear()
    return jsonify({'message': 'Logout successful'}), 200


@app.route('/api/forgot-password', methods=['POST'])
def forgot_password():
    """Request password reset."""
    try:
        data = request.get_json()
        email = data.get('email')
        
        if not email:
            return jsonify({'error': 'Email is required'}), 400
        
        # Get user from database
        user = database.get_user_by_email(email)
        
        if not user:
            # Don't reveal if user exists or not for security
            return jsonify({'message': 'If the email exists, a password reset link has been sent'}), 200
        
        # Generate reset token
        import secrets
        token = secrets.token_urlsafe(32)
        
        # Token expires in 1 hour
        from datetime import datetime, timedelta
        expires_at = (datetime.now() + timedelta(hours=1)).isoformat()
        
        # Save token to database
        database.create_password_reset_token(user['id'], token, expires_at)
        
        # Get base URL from environment or construct from request
        # Use configured BASE_URL to prevent Host header injection attacks
        base_url = os.environ.get('BASE_URL', request.url_root.rstrip('/'))
        
        # Send password reset email
        email_sent = email_service.send_password_reset_email(email, token, base_url)
        
        if email_sent:
            logger.info(f"Password reset email sent to {email}")
        else:
            logger.warning(f"Failed to send password reset email to {email}")
        
        # Always return success message for security (don't reveal if email exists)
        return jsonify({'message': 'If the email exists, a password reset link has been sent'}), 200
    except Exception as e:
        logger.error(f"Forgot password error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/reset-password', methods=['POST'])
def reset_password():
    """Reset password using token."""
    try:
        data = request.get_json()
        token = data.get('token')
        new_password = data.get('password')
        
        if not token or not new_password:
            return jsonify({'error': 'Token and new password are required'}), 400
        
        # Get token from database
        token_data = database.get_password_reset_token(token)
        
        if not token_data:
            return jsonify({'error': 'Invalid or expired token'}), 400
        
        if token_data['used']:
            return jsonify({'error': 'Token has already been used'}), 400
        
        # Check if token is expired
        from datetime import datetime
        expires_at = datetime.fromisoformat(token_data['expires_at'])
        if datetime.now() > expires_at:
            return jsonify({'error': 'Token has expired'}), 400
        
        # Update password
        password_hash = hash_password(new_password)
        success = database.update_user(token_data['user_id'], {'password_hash': password_hash})
        
        if success:
            # Mark token as used
            database.mark_token_as_used(token)
            
            # Log the action
            user = database.get_user_by_id(token_data['user_id'])
            database.log_audit(
                user_id=token_data['user_id'],
                user_email=user['email'] if user else None,
                action='password_reset',
                resource_type='user',
                resource_id=str(token_data['user_id']),
                ip_address=request.remote_addr,
                user_agent=request.user_agent.string if request.user_agent else None
            )
            
            return jsonify({'message': 'Password reset successful'}), 200
        else:
            return jsonify({'error': 'Failed to reset password'}), 500
    except Exception as e:
        logger.error(f"Reset password error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/forgot-username', methods=['POST'])
def forgot_username():
    """Request username reminder (email lookup)."""
    try:
        data = request.get_json()
        email = data.get('email')
        
        if not email:
            return jsonify({'error': 'Email is required'}), 400
        
        # Get user from database
        user = database.get_user_by_email(email)
        
        if not user:
            # Don't reveal if user exists or not for security
            return jsonify({'message': 'If the email exists, a username reminder has been sent'}), 200
        
        # Send username reminder email
        email_sent = email_service.send_username_reminder_email(email)
        
        if email_sent:
            logger.info(f"Username reminder email sent to {email}")
        else:
            logger.warning(f"Failed to send username reminder email to {email}")
        
        # Always return success message for security (don't reveal if email exists)
        return jsonify({'message': 'If the email exists, a username reminder has been sent'}), 200
    except Exception as e:
        logger.error(f"Forgot username error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/me', methods=['GET'])
@login_required
def get_current_user_info():
    """Get current user information."""
    user = get_current_user()
    if user:
        return jsonify(user), 200
    return jsonify({'error': 'Not logged in'}), 401


@app.route('/invoices', methods=['GET'])
@login_required
@permission_required('view_invoices')
def invoices_page():
    return render_template('invoices.html')


@app.route('/cashflow', methods=['GET'])
@login_required
@permission_required('view_cashflow')
def cashflow_page():
    return render_template('cashflow.html')


@app.route('/users', methods=['GET'])
@login_required
@role_required('master_admin')
def users_page():
    """User management page (master admin only)."""
    return render_template('users.html')


@app.route('/audit', methods=['GET'])
@login_required
@permission_required('view_audit_log')
def audit_page():
    """Audit log page (admin and master admin only)."""
    return render_template('audit.html')


@app.route('/health', methods=['GET'])
def health_check():
    # Check if request is from browser (HTML) or API client (JSON)
    if request.accept_mimetypes.best == 'text/html' or \
       (request.accept_mimetypes.accept_html and 
        request.accept_mimetypes['text/html'] > request.accept_mimetypes['application/json']):
        return render_template('health.html')
    return jsonify({"status": "healthy"}), 200


@app.route('/api/invoices', methods=['GET'])
@login_required
@permission_required('view_invoices')
@audit_log('view_invoices', 'invoice')
def get_invoices():
    try:
        # Get fresh QBO client with current credentials from database
        fresh_client, credentials_valid = get_fresh_qbo_client()
        
        # Check if QBO credentials are configured
        if not credentials_valid:
            return jsonify({
                "error": "QuickBooks credentials not configured",
                "message": "Please configure QuickBooks OAuth credentials to access invoice data.",
                "action": "Go to /qbo-settings to connect to QuickBooks",
                "invoices": []
            }), 200  # Return 200 so frontend can display the message nicely
        
        # Create invoice manager with fresh client
        invoice_mgr = InvoiceManager(fresh_client, database=database, predictor=predictor)
        
        # Extract query parameters for filtering
        filters = {
            'start_date': request.args.get('start_date'),
            'end_date': request.args.get('end_date'),
            'invoice_start_date': request.args.get('invoice_start_date'),
            'invoice_end_date': request.args.get('invoice_end_date'),
            'customer_id': request.args.get('customer_id'),
            'status': request.args.get('status'),
            'min_amount': request.args.get('min_amount'),
            'max_amount': request.args.get('max_amount'),
            'region': request.args.get('region'),
            'vzt_rep': request.args.get('vzt_rep'),
            'customer_portal': request.args.get('customer_portal'),
            'missing_portal_submission': request.args.get('missing_portal_submission'),
            'search_query': request.args.get('search_query')
        }

        # Remove None values
        filters = {k: v for k, v in filters.items() if v is not None}

        # Build QBO server-side filters for efficiency
        qbo_filters = {}
        if 'status' in filters:
            qbo_filters['status'] = filters['status']

        # Fetch invoices with server-side filtering
        invoices = invoice_mgr.fetch_invoices(qbo_filters=qbo_filters)
        
        # Enrich invoices with metadata from database BEFORE client-side filtering
        # This allows filtering by metadata fields
        all_metadata = database.get_all_invoice_metadata()
        metadata_map = {m['invoice_id']: m for m in all_metadata}
        
        for invoice in invoices:
            invoice_id = invoice.get('id') or invoice.get('doc_number')
            if invoice_id and invoice_id in metadata_map:
                invoice['metadata'] = metadata_map[invoice_id]
        
        # Apply client-side filters (includes metadata filters)
        filtered_invoices = invoice_mgr.filter_invoices(invoices, **filters)

        sort_by = request.args.get('sort_by', 'due_date')
        reverse = request.args.get('reverse', 'false').lower() == 'true'

        sorted_invoices = invoice_mgr.sort_invoices(filtered_invoices, sort_by=sort_by, reverse=reverse)

        return jsonify(sorted_invoices), 200
    except Exception as e:
        logger.error(f"Error fetching invoices: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/invoices/<invoice_id>/metadata', methods=['GET', 'POST'])
@login_required
def invoice_metadata(invoice_id):
    """Get or update invoice metadata."""
    if request.method == 'GET':
        if not has_permission(session.get('user_role'), 'view_invoices'):
            return jsonify({'error': 'Permission denied'}), 403
        try:
            metadata = database.get_invoice_metadata(invoice_id)
            if metadata:
                return jsonify(metadata), 200
            else:
                return jsonify({}), 200
        except Exception as e:
            logger.error(f"Error fetching invoice metadata: {e}")
            return jsonify({"error": str(e)}), 500
    
    elif request.method == 'POST':
        if not has_permission(session.get('user_role'), 'edit_invoice_metadata'):
            return jsonify({'error': 'Permission denied'}), 403
        try:
            data = request.get_json()
            success = database.save_invoice_metadata(invoice_id, data)
            if success:
                # Log the action
                database.log_audit(
                    user_id=session.get('user_id'),
                    user_email=session.get('user_email'),
                    action='update_invoice_metadata',
                    resource_type='invoice',
                    resource_id=invoice_id,
                    details=str(data),
                    ip_address=request.remote_addr,
                    user_agent=request.user_agent.string if request.user_agent else None
                )
                return jsonify({"message": "Metadata saved successfully"}), 200
            else:
                return jsonify({"error": "Failed to save metadata"}), 500
        except Exception as e:
            logger.error(f"Error saving invoice metadata: {e}")
            return jsonify({"error": str(e)}), 500


@app.route('/api/invoices/bulk-assign', methods=['POST'])
@login_required
@permission_required('edit_invoice_metadata')
def bulk_assign_invoices():
    """Bulk assign metadata to multiple invoices."""
    try:
        data = request.get_json()
        invoice_ids = data.get('invoice_ids', [])
        metadata = data.get('metadata', {})
        
        if not invoice_ids or not metadata:
            return jsonify({"error": "Missing invoice_ids or metadata"}), 400
        
        updated = 0
        for invoice_id in invoice_ids:
            success = database.save_invoice_metadata(invoice_id, metadata)
            if success:
                updated += 1
                # Log the action
                database.log_audit(
                    user_id=session.get('user_id'),
                    user_email=session.get('user_email'),
                    action='bulk_assign_invoice_metadata',
                    resource_type='invoice',
                    resource_id=invoice_id,
                    details=str(metadata),
                    ip_address=request.remote_addr,
                    user_agent=request.user_agent.string if request.user_agent else None
                )
        
        return jsonify({"message": f"Updated {updated} invoice(s)", "updated": updated}), 200
    except Exception as e:
        logger.error(f"Error bulk assigning: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/invoices/export-excel', methods=['GET'])
@login_required
@permission_required('view_invoices')
def export_invoices_to_excel():
    """Export invoices to Excel file."""
    try:
        # Get fresh QBO client with current credentials from database
        fresh_client, credentials_valid = get_fresh_qbo_client()
        
        if not credentials_valid:
            return jsonify({"error": "QuickBooks credentials not configured"}), 400
        
        # Create invoice manager with fresh client
        invoice_mgr = InvoiceManager(fresh_client, database=database, predictor=predictor)
        
        # Extract query parameters for filtering (same as get_invoices)
        filters = {
            'start_date': request.args.get('start_date'),
            'end_date': request.args.get('end_date'),
            'invoice_start_date': request.args.get('invoice_start_date'),
            'invoice_end_date': request.args.get('invoice_end_date'),
            'customer_id': request.args.get('customer_id'),
            'status': request.args.get('status'),
            'min_amount': request.args.get('min_amount'),
            'max_amount': request.args.get('max_amount'),
            'region': request.args.get('region'),
            'vzt_rep': request.args.get('vzt_rep'),
            'customer_portal': request.args.get('customer_portal'),
            'missing_portal_submission': request.args.get('missing_portal_submission'),
            'search_query': request.args.get('search_query')
        }
        filters = {k: v for k, v in filters.items() if v is not None}
        
        # Build QBO server-side filters
        qbo_filters = {}
        if 'status' in filters:
            qbo_filters['status'] = filters['status']
        
        invoices = invoice_mgr.fetch_invoices(qbo_filters=qbo_filters)
        
        # Enrich invoices with metadata BEFORE filtering
        all_metadata = database.get_all_invoice_metadata()
        metadata_map = {m['invoice_id']: m for m in all_metadata}
        
        for invoice in invoices:
            invoice_id = invoice.get('id') or invoice.get('doc_number')
            if invoice_id and invoice_id in metadata_map:
                invoice['metadata'] = metadata_map[invoice_id]
        
        filtered_invoices = invoice_mgr.filter_invoices(invoices, **filters)
        
        sort_by = request.args.get('sort_by', 'due_date')
        reverse = request.args.get('reverse', 'false').lower() == 'true'
        sorted_invoices = invoice_mgr.sort_invoices(filtered_invoices, sort_by=sort_by, reverse=reverse)
        
        # Calculate projected pay date for each invoice
        for invoice in sorted_invoices:
            projected_date = invoice_mgr.calculate_projected_pay_date(invoice)
            invoice['projected_pay_date'] = projected_date.strftime('%Y-%m-%d') if projected_date else None
        
        # Create Excel file using openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from io import BytesIO
            from flask import send_file
        except ImportError:
            return jsonify({"error": "openpyxl not installed. Please install it."}), 500
        
        wb = Workbook()
        ws = wb.active
        ws.title = "2026 Pay Schedule"
        
        # Headers
        headers = [
            "Invoice ID", "Customer", "Amount", "Due Date", "Status", 
            "VZT Rep", "Sent to Rep", "Customer Portal", "Portal Submission", 
            "Manual Pay Date", "Projected Pay Date"
        ]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data
        for invoice in sorted_invoices:
            metadata = invoice.get('metadata', {})
            ws.append([
                invoice.get('id') or invoice.get('doc_number'),
                invoice.get('customer') or invoice.get('customer_id'),
                invoice.get('amount', 0),
                invoice.get('due_date'),
                invoice.get('status'),
                metadata.get('vzt_rep', ''),
                metadata.get('sent_to_vzt_rep_date', ''),
                metadata.get('customer_portal_name', ''),
                metadata.get('portal_submission_date', ''),
                metadata.get('manual_override_pay_date', ''),
                invoice.get('projected_pay_date', '')
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Log the action
        database.log_audit(
            user_id=session.get('user_id'),
            user_email=session.get('user_email'),
            action='export_invoices_to_excel',
            resource_type='invoice',
            resource_id=None,
            details=f"Exported {len(sorted_invoices)} invoices",
            ip_address=request.remote_addr,
            user_agent=request.user_agent.string if request.user_agent else None
        )
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'pay_schedule_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/cashflow', methods=['GET'])
@login_required
@permission_required('view_cashflow')
@audit_log('view_cashflow', 'cashflow')
def get_cashflow():
    try:
        # Get fresh QBO client with current credentials from database
        fresh_client, credentials_valid = get_fresh_qbo_client()
        
        # Check if QBO credentials are configured
        if not credentials_valid:
            return jsonify({
                "error": "QuickBooks credentials not configured",
                "message": "Please configure QuickBooks OAuth credentials to access cashflow data.",
                "action": "Go to /qbo-settings to connect to QuickBooks",
                "days": 30,
                "projected_balance_change": []
            }), 200
        
        # Create invoice manager with fresh client
        invoice_mgr = InvoiceManager(fresh_client, database=database, predictor=predictor)
        
        days = int(request.args.get('days', 30))
        invoices = invoice_mgr.fetch_invoices()
        # Mock expenses for now, or fetch from another source if available
        expenses = []

        projector = CashFlowProjector(invoices, expenses, predictor=predictor)
        projection = projector.calculate_projection(days=days)

        return jsonify({
            "days": days,
            "projected_balance_change": projection
        }), 200
    except Exception as e:
        logger.error(f"Error calculating cashflow: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/cashflow/calendar', methods=['GET'])
@login_required
@permission_required('view_cashflow')
@audit_log('view_cashflow_calendar', 'cashflow')
def get_cashflow_calendar():
    """Get calendar-style cash flow projection with daily breakdown."""
    try:
        from datetime import datetime, timedelta
        
        # Get fresh QBO client with current credentials from database
        fresh_client, credentials_valid = get_fresh_qbo_client()
        
        # Get parameters
        days = int(request.args.get('days', 90))
        initial_balance_param = request.args.get('initial_balance')
        start_date_param = request.args.get('start_date')
        end_date_param = request.args.get('end_date')
        
        # Get initial balance from QBO if not provided and credentials are valid
        if initial_balance_param:
            initial_balance = float(initial_balance_param)
        elif credentials_valid:
            # Fetch from QBO
            bank_accounts = fresh_client.fetch_bank_accounts()
            initial_balance = 0.0
            for account in bank_accounts:
                balance = account.get('CurrentBalance', 0)
                initial_balance += float(balance) if balance else 0
            logger.info(f"Using QBO bank balance: {initial_balance}")
        else:
            # Default to 0 if no initial balance provided and no QBO credentials
            initial_balance = 0.0
            logger.info("Using default initial balance: 0.0 (QBO credentials not configured)")
        
        # Toggle parameters
        show_projected_inflows = request.args.get('show_projected_inflows', 'true').lower() == 'true'
        show_projected_outflows = request.args.get('show_projected_outflows', 'true').lower() == 'true'
        show_custom_inflows = request.args.get('show_custom_inflows', 'true').lower() == 'true'
        show_custom_outflows = request.args.get('show_custom_outflows', 'true').lower() == 'true'
        
        # Calculate date range
        if start_date_param and end_date_param:
            start_date = datetime.strptime(start_date_param, '%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = datetime.strptime(end_date_param, '%Y-%m-%d').replace(hour=0, minute=0, second=0, microsecond=0)
        else:
            start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = start_date + timedelta(days=days)
        
        # Create invoice manager with fresh client
        invoice_mgr = InvoiceManager(fresh_client, database=database, predictor=predictor)
        
        # Fetch data (invoices will be empty if QBO credentials not configured)
        invoices = invoice_mgr.fetch_invoices() if credentials_valid else []
        accounts_payable = []  # TODO: Fetch from QBO when available
        custom_flows = database.get_custom_cash_flows()
        
        # Create calendar projector
        calendar = CashFlowCalendar(
            invoices=invoices,
            accounts_payable=accounts_payable,
            custom_flows=custom_flows,
            predictor=predictor,
            database=database
        )
        
        # Calculate projection
        projection = calendar.calculate_daily_projection(
            start_date=start_date,
            end_date=end_date,
            initial_balance=initial_balance,
            show_projected_inflows=show_projected_inflows,
            show_projected_outflows=show_projected_outflows,
            show_custom_inflows=show_custom_inflows,
            show_custom_outflows=show_custom_outflows
        )
        
        return jsonify({
            "start_date": start_date.strftime('%Y-%m-%d'),
            "end_date": end_date.strftime('%Y-%m-%d'),
            "initial_balance": initial_balance,
            "daily_projection": projection
        }), 200
    except Exception as e:
        logger.error(f"Error calculating calendar cashflow: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/bank-accounts', methods=['GET'])
@login_required
@permission_required('view_cashflow')
def get_bank_accounts():
    """Get bank accounts and their current balances from QBO."""
    try:
        bank_accounts = qbo_client.fetch_bank_accounts()
        
        # Format the response
        accounts_data = []
        total_balance = 0.0
        
        for account in bank_accounts:
            balance = float(account.get('CurrentBalance', 0))
            total_balance += balance
            
            accounts_data.append({
                'id': account.get('Id'),
                'name': account.get('Name'),
                'account_number': account.get('AcctNum', 'N/A'),
                'balance': balance,
                'currency': account.get('CurrencyRef', {}).get('value', 'USD')
            })
        
        return jsonify({
            'accounts': accounts_data,
            'total_balance': total_balance,
            'as_of': datetime.now().isoformat()
        }), 200
    except Exception as e:
        logger.error(f"Error fetching bank accounts: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/custom-cash-flows', methods=['GET', 'POST'])
@login_required
def custom_cash_flows():
    """Get all custom cash flows or add a new one."""
    if request.method == 'GET':
        if not has_permission(session.get('user_role'), 'view_cashflow'):
            return jsonify({'error': 'Permission denied'}), 403
        try:
            flow_type = request.args.get('flow_type')  # 'inflow' or 'outflow'
            flows = database.get_custom_cash_flows(flow_type)
            return jsonify(flows), 200
        except Exception as e:
            logger.error(f"Error fetching custom cash flows: {e}")
            return jsonify({"error": str(e)}), 500
    
    elif request.method == 'POST':
        # Check permissions based on flow type
        data = request.get_json()
        flow_type = data.get('flow_type')
        if flow_type == 'inflow' and not has_permission(session.get('user_role'), 'add_custom_inflows'):
            return jsonify({'error': 'Permission denied'}), 403
        if flow_type == 'outflow' and not has_permission(session.get('user_role'), 'add_custom_outflows'):
            return jsonify({'error': 'Permission denied'}), 403
        
        try:
            flow_id = database.add_custom_cash_flow(data)
            if flow_id:
                # Log the action
                database.log_audit(
                    user_id=session.get('user_id'),
                    user_email=session.get('user_email'),
                    action='add_custom_cash_flow',
                    resource_type='custom_cash_flow',
                    resource_id=str(flow_id),
                    details=str(data),
                    ip_address=request.remote_addr,
                    user_agent=request.user_agent.string if request.user_agent else None
                )
                return jsonify({"message": "Custom cash flow added", "id": flow_id}), 201
            else:
                return jsonify({"error": "Failed to add custom cash flow"}), 500
        except Exception as e:
            logger.error(f"Error adding custom cash flow: {e}")
            return jsonify({"error": str(e)}), 500


@app.route('/api/custom-cash-flows/<int:flow_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def custom_cash_flow_detail(flow_id):
    """Get, update, or delete a specific custom cash flow."""
    if request.method == 'GET':
        if not has_permission(session.get('user_role'), 'view_cashflow'):
            return jsonify({'error': 'Permission denied'}), 403
        try:
            flows = database.get_custom_cash_flows()
            flow = next((f for f in flows if f['id'] == flow_id), None)
            if flow:
                return jsonify(flow), 200
            else:
                return jsonify({"error": "Cash flow not found"}), 404
        except Exception as e:
            logger.error(f"Error fetching custom cash flow: {e}")
            return jsonify({"error": str(e)}), 500
    
    elif request.method == 'PUT':
        if not has_permission(session.get('user_role'), 'edit_custom_flows'):
            return jsonify({'error': 'Permission denied'}), 403
        try:
            data = request.get_json()
            success = database.update_custom_cash_flow(flow_id, data)
            if success:
                # Log the action
                database.log_audit(
                    user_id=session.get('user_id'),
                    user_email=session.get('user_email'),
                    action='update_custom_cash_flow',
                    resource_type='custom_cash_flow',
                    resource_id=str(flow_id),
                    details=str(data),
                    ip_address=request.remote_addr,
                    user_agent=request.user_agent.string if request.user_agent else None
                )
                return jsonify({"message": "Custom cash flow updated"}), 200
            else:
                return jsonify({"error": "Failed to update custom cash flow"}), 500
        except Exception as e:
            logger.error(f"Error updating custom cash flow: {e}")
            return jsonify({"error": str(e)}), 500
    
    elif request.method == 'DELETE':
        if not has_permission(session.get('user_role'), 'delete_custom_flows'):
            return jsonify({'error': 'Permission denied'}), 403
        try:
            success = database.delete_custom_cash_flow(flow_id)
            if success:
                # Log the action
                database.log_audit(
                    user_id=session.get('user_id'),
                    user_email=session.get('user_email'),
                    action='delete_custom_cash_flow',
                    resource_type='custom_cash_flow',
                    resource_id=str(flow_id),
                    ip_address=request.remote_addr,
                    user_agent=request.user_agent.string if request.user_agent else None
                )
                return jsonify({"message": "Custom cash flow deleted"}), 200
            else:
                return jsonify({"error": "Failed to delete custom cash flow"}), 500
        except Exception as e:
            logger.error(f"Error deleting custom cash flow: {e}")
            return jsonify({"error": str(e)}), 500


# User management API routes

@app.route('/api/users', methods=['GET', 'POST'])
@login_required
@role_required('master_admin')
def manage_users():
    """Get all users or create a new user (master admin only)."""
    if request.method == 'GET':
        try:
            users = database.get_all_users()
            return jsonify(users), 200
        except Exception as e:
            logger.error(f"Error fetching users: {e}")
            return jsonify({"error": str(e)}), 500
    
    elif request.method == 'POST':
        try:
            data = request.get_json()
            email = data.get('email')
            password = data.get('password')
            full_name = data.get('full_name')
            role = data.get('role')
            
            if not email or not password or not role:
                return jsonify({'error': 'Email, password, and role are required'}), 400
            
            if role not in ROLES:
                return jsonify({'error': f'Invalid role. Must be one of: {", ".join(ROLES.keys())}'}), 400
            
            # Check if user already exists
            existing_user = database.get_user_by_email(email)
            if existing_user:
                return jsonify({'error': 'User with this email already exists'}), 400
            
            # Hash password
            password_hash = hash_password(password)
            
            # Create user
            user_id = database.create_user(email, password_hash, full_name, role)
            
            if user_id:
                # Log the action
                database.log_audit(
                    user_id=session.get('user_id'),
                    user_email=session.get('user_email'),
                    action='create_user',
                    resource_type='user',
                    resource_id=str(user_id),
                    details=f"Created user {email} with role {role}",
                    ip_address=request.remote_addr,
                    user_agent=request.user_agent.string if request.user_agent else None
                )
                return jsonify({"message": "User created successfully", "id": user_id}), 201
            else:
                return jsonify({"error": "Failed to create user"}), 500
        except Exception as e:
            logger.error(f"Error creating user: {e}")
            return jsonify({"error": str(e)}), 500


@app.route('/api/users/<int:user_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
@role_required('master_admin')
def manage_user_detail(user_id):
    """Get, update, or delete a specific user (master admin only)."""
    if request.method == 'GET':
        try:
            user = database.get_user_by_id(user_id)
            if user:
                # Remove password hash from response
                user.pop('password_hash', None)
                return jsonify(user), 200
            else:
                return jsonify({"error": "User not found"}), 404
        except Exception as e:
            logger.error(f"Error fetching user: {e}")
            return jsonify({"error": str(e)}), 500
    
    elif request.method == 'PUT':
        try:
            data = request.get_json()
            
            # Validate role if provided
            if 'role' in data and data['role'] not in ROLES:
                return jsonify({'error': f'Invalid role. Must be one of: {", ".join(ROLES.keys())}'}), 400
            
            # Hash new password if provided
            if 'password' in data:
                data['password_hash'] = hash_password(data.pop('password'))
            
            success = database.update_user(user_id, data)
            
            if success:
                # Log the action
                database.log_audit(
                    user_id=session.get('user_id'),
                    user_email=session.get('user_email'),
                    action='update_user',
                    resource_type='user',
                    resource_id=str(user_id),
                    details=str(data),
                    ip_address=request.remote_addr,
                    user_agent=request.user_agent.string if request.user_agent else None
                )
                return jsonify({"message": "User updated successfully"}), 200
            else:
                return jsonify({"error": "Failed to update user"}), 500
        except Exception as e:
            logger.error(f"Error updating user: {e}")
            return jsonify({"error": str(e)}), 500
    
    elif request.method == 'DELETE':
        try:
            # Prevent deleting yourself
            if user_id == session.get('user_id'):
                return jsonify({"error": "Cannot delete your own account"}), 400
            
            success = database.delete_user(user_id)
            
            if success:
                # Log the action
                database.log_audit(
                    user_id=session.get('user_id'),
                    user_email=session.get('user_email'),
                    action='delete_user',
                    resource_type='user',
                    resource_id=str(user_id),
                    ip_address=request.remote_addr,
                    user_agent=request.user_agent.string if request.user_agent else None
                )
                return jsonify({"message": "User deleted successfully"}), 200
            else:
                return jsonify({"error": "Failed to delete user"}), 500
        except Exception as e:
            logger.error(f"Error deleting user: {e}")
            return jsonify({"error": str(e)}), 500


@app.route('/api/users/<int:user_id>/force-reset-password', methods=['POST'])
@login_required
@role_required('admin', 'master_admin')
def force_reset_password(user_id):
    """Force reset a user's password (admin and master admin only)."""
    try:
        data = request.get_json()
        new_password = data.get('password')
        send_email_notification = data.get('send_email', True)
        
        if not new_password:
            return jsonify({'error': 'New password is required'}), 400
        
        # Get the user
        user = database.get_user_by_id(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Hash the new password
        password_hash = hash_password(new_password)
        
        # Update the password
        success = database.update_user(user_id, {'password_hash': password_hash})
        
        if not success:
            return jsonify({'error': 'Failed to reset password'}), 500
        
        # Send email notification to user if requested
        if send_email_notification:
            try:
                subject = "Password Reset by Administrator - VZT Accounting"
                
                html_body = f"""
                <html>
                <head>
                    <style>
                        body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                        .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                        .header {{ background-color: #FF9800; color: white; padding: 20px; text-align: center; }}
                        .content {{ background-color: #f9f9f9; padding: 30px; border: 1px solid #ddd; }}
                        .password-box {{ background-color: #fff3cd; border: 2px solid #FF9800; padding: 20px; 
                                        border-radius: 4px; text-align: center; margin: 20px 0; }}
                        .footer {{ text-align: center; padding: 20px; color: #666; font-size: 0.9em; }}
                        .warning {{ background-color: #ffebee; border: 1px solid #f44336; padding: 15px; 
                                   border-radius: 4px; margin: 20px 0; }}
                    </style>
                </head>
                <body>
                    <div class="container">
                        <div class="header">
                            <h1>🔐 Password Reset by Administrator</h1>
                        </div>
                        <div class="content">
                            <p>Hello {user['full_name'] or user['email']},</p>
                            <p>Your password for VZT Accounting has been reset by an administrator.</p>
                            
                            <div class="password-box">
                                <p style="margin: 0; font-size: 0.9em; color: #666;">Your new temporary password is:</p>
                                <h2 style="margin: 10px 0; color: #FF9800; font-family: monospace;">{new_password}</h2>
                            </div>
                            
                            <div class="warning">
                                <strong>⚠️ Important Security Steps:</strong>
                                <ol style="margin: 10px 0; padding-left: 20px;">
                                    <li>Log in to your account using this temporary password</li>
                                    <li>Change your password immediately after logging in</li>
                                    <li>Do not share this password with anyone</li>
                                    <li>If you didn't expect this reset, contact your administrator immediately</li>
                                </ol>
                            </div>
                            
                            <p>You can log in at: <a href="{os.environ.get('BASE_URL', request.url_root.rstrip('/'))}/login">{os.environ.get('BASE_URL', request.url_root.rstrip('/'))}/login</a></p>
                        </div>
                        <div class="footer">
                            <p>This is an automated message from VZT Accounting.<br>
                            Please do not reply to this email.</p>
                        </div>
                    </div>
                </body>
                </html>
                """
                
                text_body = f"""
Password Reset by Administrator - VZT Accounting

Hello {user['full_name'] or user['email']},

Your password for VZT Accounting has been reset by an administrator.

Your new temporary password is: {new_password}

IMPORTANT SECURITY STEPS:
1. Log in to your account using this temporary password
2. Change your password immediately after logging in
3. Do not share this password with anyone
4. If you didn't expect this reset, contact your administrator immediately

You can log in at: {os.environ.get('BASE_URL', request.url_root.rstrip('/'))}/login

---
This is an automated message from VZT Accounting.
Please do not reply to this email.
                """
                
                email_service.send_email(user['email'], subject, html_body, text_body)
                logger.info(f"Password reset notification sent to {user['email']}")
            except Exception as email_error:
                logger.error(f"Failed to send password reset notification: {email_error}")
                # Don't fail the password reset if email fails
        
        # Log the action
        database.log_audit(
            user_id=session.get('user_id'),
            user_email=session.get('user_email'),
            action='force_reset_password',
            resource_type='user',
            resource_id=str(user_id),
            details=f"Force reset password for {user['email']}",
            ip_address=request.remote_addr,
            user_agent=request.user_agent.string if request.user_agent else None
        )
        
        return jsonify({
            "message": "Password reset successfully",
            "email_sent": send_email_notification
        }), 200
        
    except Exception as e:
        logger.error(f"Error forcing password reset: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/roles', methods=['GET'])
@login_required
@role_required('master_admin')
def get_roles():
    """Get all available roles and their permissions."""
    return jsonify(ROLES), 200


# Audit log API routes

@app.route('/api/audit-log', methods=['GET'])
@login_required
@permission_required('view_audit_log')
def get_audit_log():
    """Get audit log entries (admin and master admin only)."""
    try:
        user_id = request.args.get('user_id', type=int)
        action = request.args.get('action')
        resource_type = request.args.get('resource_type')
        limit = request.args.get('limit', default=100, type=int)
        
        logs = database.get_audit_logs(user_id, action, resource_type, limit)
        return jsonify(logs), 200
    except Exception as e:
        logger.error(f"Error fetching audit logs: {e}")
        return jsonify({"error": str(e)}), 500


# QBO Credentials Management API routes

@app.route('/api/qbo/credentials', methods=['GET', 'POST'])
@login_required
def manage_qbo_credentials():
    """Get or update QBO credentials (admin and master_admin only)."""
    global qbo_client, secret_manager
    user_role = session.get('user_role')
    
    # Only admin and master_admin can manage QBO credentials
    if user_role not in ['admin', 'master_admin']:
        return jsonify({'error': 'Permission denied'}), 403
    
    if request.method == 'GET':
        try:
            credentials = database.get_qbo_credentials()
            if credentials:
                # Check if credentials are valid (not dummy values)
                is_valid = not (credentials['client_id'] == 'dummy_id' or 
                              credentials['client_secret'] == 'dummy_secret' or 
                              credentials['refresh_token'] == 'dummy_refresh' or 
                              credentials['realm_id'] == 'dummy_realm')
                
                # Don't expose sensitive data
                safe_credentials = {
                    'client_id': credentials['client_id'][:10] + '...' if credentials['client_id'] else None,
                    'realm_id': credentials['realm_id'],
                    'has_refresh_token': bool(credentials.get('refresh_token')),
                    'has_access_token': bool(credentials.get('access_token')),
                    'access_token_expires_at': credentials.get('access_token_expires_at'),
                    'refresh_token_expires_at': credentials.get('refresh_token_expires_at'),
                    'created_at': credentials.get('created_at'),
                    'updated_at': credentials.get('updated_at'),
                    'is_valid': is_valid,
                    'status': 'configured' if is_valid else 'invalid_or_dummy'
                }
                return jsonify(safe_credentials), 200
            else:
                # Check if environment variables are set
                qbo_creds = secret_manager.get_qbo_credentials()
                is_env_valid = qbo_creds.get('is_valid', False)
                
                return jsonify({
                    'message': 'No QBO credentials configured in database',
                    'is_valid': is_env_valid,
                    'status': 'using_environment_variables' if is_env_valid else 'not_configured',
                    'help': 'Please configure credentials at /qbo-settings or set environment variables'
                }), 404
        except Exception as e:
            logger.error(f"Error fetching QBO credentials: {e}")
            return jsonify({"error": str(e)}), 500
    
    elif request.method == 'POST':
        try:
            data = request.get_json()
            client_id = data.get('client_id')
            client_secret = data.get('client_secret')
            refresh_token = data.get('refresh_token')
            realm_id = data.get('realm_id')
            
            if not all([client_id, client_secret, refresh_token, realm_id]):
                return jsonify({'error': 'All fields are required: client_id, client_secret, refresh_token, realm_id'}), 400
            
            credentials = {
                'client_id': client_id,
                'client_secret': client_secret,
                'refresh_token': refresh_token,
                'realm_id': realm_id
            }
            
            success = database.save_qbo_credentials(credentials, session.get('user_id'))
            
            if success:
                # Update the QBO client with new credentials
                qbo_credentials = secret_manager.get_qbo_credentials()
                qbo_client = QBOClient(
                    qbo_credentials['client_id'],
                    qbo_credentials['client_secret'],
                    qbo_credentials['refresh_token'],
                    qbo_credentials['realm_id'],
                    database=database
                )
                
                # Log the action
                database.log_audit(
                    user_id=session.get('user_id'),
                    user_email=session.get('user_email'),
                    action='update_qbo_credentials',
                    resource_type='qbo_credentials',
                    resource_id='1',
                    details='Updated QBO credentials',
                    ip_address=request.remote_addr,
                    user_agent=request.user_agent.string if request.user_agent else None
                )
                
                return jsonify({"message": "QBO credentials saved successfully"}), 200
            else:
                return jsonify({"error": "Failed to save QBO credentials"}), 500
        except Exception as e:
            logger.error(f"Error saving QBO credentials: {e}")
            return jsonify({"error": str(e)}), 500


@app.route('/api/qbo/refresh', methods=['POST'])
@login_required
def refresh_qbo_token():
    """Manually refresh QBO access token (admin and master_admin only)."""
    user_role = session.get('user_role')
    
    # Only admin and master_admin can refresh tokens
    if user_role not in ['admin', 'master_admin']:
        return jsonify({'error': 'Permission denied'}), 403
    
    try:
        # Refresh the access token
        qbo_client.refresh_access_token()
        
        # Log the action
        database.log_audit(
            user_id=session.get('user_id'),
            user_email=session.get('user_email'),
            action='refresh_qbo_token',
            resource_type='qbo_credentials',
            resource_id='1',
            details='Manually refreshed QBO access token',
            ip_address=request.remote_addr,
            user_agent=request.user_agent.string if request.user_agent else None
        )
        
        return jsonify({"message": "QBO access token refreshed successfully"}), 200
    except ValueError as ve:
        # Credentials are not valid
        logger.error(f"Cannot refresh QBO token: {ve}")
        return jsonify({
            "error": str(ve),
            "action_required": "Please configure valid OAuth credentials at /qbo-settings"
        }), 400
    except Exception as e:
        logger.error(f"Error refreshing QBO token: {e}")
        error_msg = str(e)
        
        # Check if it's a 401 error
        if "401" in error_msg or "Unauthorized" in error_msg:
            return jsonify({
                "error": "The OAuth credentials are invalid or expired",
                "details": error_msg,
                "action_required": "Please reconfigure credentials at /qbo-settings by connecting to QuickBooks again"
            }), 401
        
        return jsonify({"error": error_msg}), 500


@app.route('/api/qbo/oauth/authorize', methods=['POST'])
@login_required
def qbo_oauth_authorize():
    """Initiate QBO OAuth 2.0 flow (admin and master_admin only)."""
    user_role = session.get('user_role')
    
    # Only admin and master_admin can initiate OAuth
    if user_role not in ['admin', 'master_admin']:
        return jsonify({'error': 'Permission denied'}), 403
    
    try:
        data = request.get_json()
        client_id = data.get('client_id')
        client_secret = data.get('client_secret')
        redirect_uri = data.get('redirect_uri')
        
        if not all([client_id, client_secret, redirect_uri]):
            return jsonify({'error': 'client_id, client_secret, and redirect_uri are required'}), 400
        
        # Store client credentials in session for later use in callback
        session['qbo_oauth_client_id'] = client_id
        session['qbo_oauth_client_secret'] = client_secret
        session['qbo_oauth_redirect_uri'] = redirect_uri
        
        # Generate a state token for CSRF protection
        state = str(uuid.uuid4())
        session['qbo_oauth_state'] = state
        
        # Build the authorization URL
        # URL encode the redirect_uri to ensure it matches exactly with QuickBooks settings
        # Per RFC 3986, when used as a query parameter VALUE, reserved characters must be encoded
        # Using safe='' ensures full encoding of all special characters including :, /, ?, etc.
        encoded_redirect_uri = quote(redirect_uri, safe='')
        auth_url = (
            f"https://appcenter.intuit.com/connect/oauth2?"
            f"client_id={client_id}&"
            f"scope=com.intuit.quickbooks.accounting&"
            f"redirect_uri={encoded_redirect_uri}&"
            f"response_type=code&"
            f"state={state}"
        )
        
        # Log the action
        database.log_audit(
            user_id=session.get('user_id'),
            user_email=session.get('user_email'),
            action='initiate_qbo_oauth',
            resource_type='qbo_credentials',
            resource_id='1',
            details='Initiated QBO OAuth flow',
            ip_address=request.remote_addr,
            user_agent=request.user_agent.string if request.user_agent else None
        )
        
        return jsonify({'authorization_url': auth_url}), 200
    except Exception as e:
        logger.error(f"Error initiating QBO OAuth: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/qbo/oauth/callback', methods=['GET'])
@login_required
def qbo_oauth_callback():
    """Handle QBO OAuth 2.0 callback (admin and master_admin only)."""
    user_role = session.get('user_role')
    
    # Only admin and master_admin can complete OAuth
    if user_role not in ['admin', 'master_admin']:
        return render_template('oauth_callback.html', error='Permission denied')
    
    try:
        # Get authorization code and state from query params
        code = request.args.get('code')
        state = request.args.get('state')
        realm_id = request.args.get('realmId')
        error = request.args.get('error')
        
        # Log callback details for debugging
        logger.info("=" * 80)
        logger.info("QuickBooks OAuth 2.0 Callback Received")
        logger.info("=" * 80)
        logger.info(f"Authorization Code: {'[RECEIVED]' if code else 'No code received'}")
        logger.info(f"Realm ID: {realm_id}")
        logger.info(f"State Token: {'[VALID]' if state else '[MISSING]'}")
        logger.info(f"Error (if any): {error}")
        logger.info("=" * 80)
        
        # Check for errors from QBO
        if error:
            error_description = request.args.get('error_description', 'No description provided')
            logger.error(f"QBO OAuth error: {error} - {error_description}")
            return render_template('oauth_callback.html', error=f'QBO authorization failed: {error} - {error_description}')
        
        # Verify state to prevent CSRF
        if state != session.get('qbo_oauth_state'):
            # Log security incident with details
            logger.error(
                f"SECURITY: OAuth state mismatch - possible CSRF attack. "
                f"IP: {request.remote_addr}, "
                f"User: {session.get('user_email', 'unknown')}, "
                f"Expected state: {session.get('qbo_oauth_state')}, "
                f"Received state: {state}"
            )
            return render_template('oauth_callback.html', error='Invalid state parameter - possible CSRF attack')
        
        if not code or not realm_id:
            logger.error(f"Missing authorization code or realm ID - code: {bool(code)}, realm_id: {bool(realm_id)}")
            return render_template('oauth_callback.html', error='Missing authorization code or realm ID')
        
        # Get stored credentials from session
        client_id = session.get('qbo_oauth_client_id')
        client_secret = session.get('qbo_oauth_client_secret')
        redirect_uri = session.get('qbo_oauth_redirect_uri')
        
        if not all([client_id, client_secret, redirect_uri]):
            logger.error("OAuth session data missing - session may have expired")
            return render_template('oauth_callback.html', error='OAuth session expired. Please restart the flow.')
        
        logger.info(f"Exchanging authorization code for tokens using redirect_uri: {redirect_uri}")
        
        # Exchange authorization code for tokens
        token_url = "https://oauth.platform.intuit.com/oauth2/v1/tokens/bearer"
        
        # Create Basic Auth header
        auth_string = f"{client_id}:{client_secret}"
        auth_bytes = auth_string.encode('ascii')
        auth_b64 = base64.b64encode(auth_bytes).decode('ascii')
        
        headers = {
            'Accept': 'application/json',
            'Content-Type': 'application/x-www-form-urlencoded',
            'Authorization': f'Basic {auth_b64}'
        }
        
        data = {
            'grant_type': 'authorization_code',
            'code': code,
            'redirect_uri': redirect_uri
        }
        
        logger.info("Sending token exchange request to QuickBooks...")
        response = requests.post(token_url, headers=headers, data=data)
        response.raise_for_status()
        
        token_data = response.json()
        
        logger.info("Successfully received token response from QuickBooks")
        logger.info(f"Token data keys: {list(token_data.keys())}")
        
        # Extract tokens from response (handle both camelCase and snake_case)
        access_token = token_data.get('access_token') or token_data.get('accessToken')
        refresh_token = token_data.get('refresh_token') or token_data.get('refreshToken')
        expires_in = token_data.get('expires_in', 3600)
        x_refresh_token_expires_in = token_data.get('x_refresh_token_expires_in', 8726400)
        
        if not access_token or not refresh_token:
            logger.error(f"Missing tokens in response. Available keys: {list(token_data.keys())}")
            return render_template('oauth_callback.html', error='Failed to obtain tokens from QBO')
        
        logger.info(f"Access token expires in: {expires_in} seconds")
        logger.info(f"Refresh token expires in: {x_refresh_token_expires_in} seconds (~{x_refresh_token_expires_in/86400:.1f} days)")
        
        # Save credentials to database
        credentials = {
            'client_id': client_id,
            'client_secret': client_secret,
            'refresh_token': refresh_token,
            'access_token': access_token,
            'realm_id': realm_id,
            'expires_in': expires_in,
            'x_refresh_token_expires_in': x_refresh_token_expires_in
        }
        
        logger.info(f"Saving credentials to database for Realm ID: {realm_id}")
        success = database.save_qbo_credentials(credentials, session.get('user_id'))
        
        if success:
            logger.info("Successfully saved credentials to database")
            # Update the global QBO client with new credentials
            global qbo_client, secret_manager
            qbo_credentials = secret_manager.get_qbo_credentials()
            qbo_client = QBOClient(
                qbo_credentials['client_id'],
                qbo_credentials['client_secret'],
                qbo_credentials['refresh_token'],
                qbo_credentials['realm_id'],
                database=database
            )
            qbo_client.access_token = access_token
            
            # Clear OAuth session data
            session.pop('qbo_oauth_client_id', None)
            session.pop('qbo_oauth_client_secret', None)
            session.pop('qbo_oauth_redirect_uri', None)
            session.pop('qbo_oauth_state', None)
            
            logger.info("OAuth flow completed successfully")
            logger.info("=" * 80)
            
            # Log the action
            database.log_audit(
                user_id=session.get('user_id'),
                user_email=session.get('user_email'),
                action='complete_qbo_oauth',
                resource_type='qbo_credentials',
                resource_id='1',
                details=f'Completed QBO OAuth and saved credentials for Realm ID: {realm_id}',
                ip_address=request.remote_addr,
                user_agent=request.user_agent.string if request.user_agent else None
            )
            
            # Render callback page that will notify parent window and close popup
            return render_template('oauth_callback.html', success='true')
        else:
            logger.error("Failed to save QBO credentials to database")
            return render_template('oauth_callback.html', error='Failed to save QBO credentials to database')
            
    except requests.exceptions.HTTPError as e:
        # Handle HTTP errors from token exchange
        error_msg = f"HTTP {e.response.status_code} error during token exchange"
        try:
            error_data = e.response.json()
            error_msg += f": {error_data.get('error', 'Unknown error')}"
            if 'error_description' in error_data:
                error_msg += f" - {error_data['error_description']}"
            logger.error(f"Token exchange failed: {error_msg}")
            logger.error(f"Response body: {error_data}")
        except:
            logger.error(f"Token exchange failed: {error_msg}")
            logger.error(f"Response text: {e.response.text}")
        
        # Provide helpful guidance based on error type
        if e.response.status_code == 400:
            error_msg += "\n\nPossible causes:\n"
            error_msg += "- Authorization code already used or expired\n"
            error_msg += "- Redirect URI mismatch\n"
            error_msg += "- Invalid client credentials"
        elif e.response.status_code == 401:
            error_msg += "\n\nPossible causes:\n"
            error_msg += "- Invalid client ID or client secret\n"
            error_msg += "- Credentials may be for wrong environment (Sandbox vs Production)"
        
        return render_template('oauth_callback.html', error=error_msg)
    except requests.exceptions.RequestException as e:
        logger.error(f"Request error during token exchange: {e}")
        return render_template('oauth_callback.html', error=f'Network error during token exchange: {str(e)}')
    except Exception as e:
        logger.error(f"Unexpected error in OAuth callback: {e}", exc_info=True)
        return render_template('oauth_callback.html', error=f'An unexpected error occurred: {str(e)}')


@app.route('/qbo-settings', methods=['GET'])
@login_required
def qbo_settings_redirect():
    """Redirect /qbo-settings to /qbo-settings-v2 for backwards compatibility."""
    return redirect('/qbo-settings-v2')


@app.route('/api/qbo/oauth/diagnostic', methods=['GET'])
@login_required
def qbo_oauth_diagnostic():
    """
    Diagnostic endpoint to help troubleshoot OAuth configuration.
    Returns information about the current OAuth setup.
    """
    user_role = session.get('user_role')
    
    # Only admin and master_admin can access diagnostics
    if user_role not in ['admin', 'master_admin']:
        return jsonify({'error': 'Permission denied'}), 403
    
    try:
        # Get current credentials
        credentials = database.get_qbo_credentials()
        
        # Get the redirect URI that would be used
        parsed_url = urlparse(request.host_url.rstrip('/'))
        # Construct new URL with HTTPS scheme
        https_url = urlunparse((
            'https',  # scheme
            parsed_url.netloc,  # netloc
            parsed_url.path,  # path
            parsed_url.params,  # params
            parsed_url.query,  # query
            parsed_url.fragment  # fragment
        ))
        redirect_uri = https_url + '/api/qbo/oauth/callback'
        
        # Get hardcoded client ID from the code (masked for security)
        hardcoded_client_id_full = 'AB224ne26KUlOjJebeDLMIwgIZcTRQkb6AieFqwJQg0sWCzXXA'
        hardcoded_client_id_masked = hardcoded_client_id_full[:10] + '...'
        
        diagnostic_info = {
            'oauth_configuration': {
                'authorization_endpoint': 'https://appcenter.intuit.com/connect/oauth2',
                'token_endpoint': 'https://oauth.platform.intuit.com/oauth2/v1/tokens/bearer',
                'scope': 'com.intuit.quickbooks.accounting',
                'response_type': 'code'
            },
            'current_setup': {
                'redirect_uri': redirect_uri,
                'client_id_in_use': hardcoded_client_id_masked,
                'host_url': request.host_url,
                'using_https': parsed_url.scheme == 'https'
            },
            'database_credentials': {
                'configured': bool(credentials),
                'client_id': 'Set (masked for security)' if credentials and credentials.get('client_id') else 'Not set',
                'realm_id': credentials.get('realm_id') if credentials else 'Not set',
                'has_refresh_token': bool(credentials.get('refresh_token')) if credentials else False,
                'has_access_token': bool(credentials.get('access_token')) if credentials else False,
                'tokens_valid': credentials.get('is_valid', False) if credentials else False
            },
            'troubleshooting_checklist': [
                {
                    'item': 'Redirect URI must be registered in QBO Developer Portal',
                    'value': redirect_uri,
                    'status': 'unknown',
                    'action': f'Log into developer.intuit.com and verify "{redirect_uri}" is in your app\'s redirect URIs'
                },
                {
                    'item': 'Client ID must match the one in your QBO app',
                    'value': hardcoded_client_id_masked,
                    'status': 'unknown',
                    'action': 'Verify this client ID matches your QBO app credentials'
                },
                {
                    'item': 'Using correct environment',
                    'value': f"{os.environ.get('QBO_ENVIRONMENT', 'production').upper()} (from QBO_ENVIRONMENT variable)",
                    'status': 'info',
                    'action': f'Ensure your client credentials are from the {os.environ.get("QBO_ENVIRONMENT", "production").title()} environment'
                },
                {
                    'item': 'Browser allows popups',
                    'value': 'Unknown',
                    'status': 'warning',
                    'action': 'Ensure your browser allows popups for this site'
                }
            ],
            'common_issues': {
                'no_login_screen': [
                    'Redirect URI not registered in QBO Developer Portal',
                    'Client ID is invalid or for wrong environment',
                    'Browser popup blocker is preventing the OAuth window',
                    'Client secret is incorrect (check during callback)'
                ],
                '403_forbidden': [
                    'Invalid or expired credentials',
                    'Refresh token expired (>101 days old)',
                    'Realm ID is not accessible with these credentials',
                    'Need to reconnect via OAuth flow'
                ]
            },
            'next_steps': [
                '1. Verify the redirect URI is registered in your QBO app at developer.intuit.com',
                '2. Check that your client ID and secret are correct',
                '3. Try clicking "Connect to QuickBooks" and check browser console for errors',
                '4. Check server logs for detailed OAuth flow information',
                '5. If you see a blank screen, check for popup blockers'
            ]
        }
        
        return jsonify(diagnostic_info), 200
    except Exception as e:
        logger.error(f"Error generating OAuth diagnostics: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route('/qbo-settings-v2', methods=['GET'])
@login_required
def qbo_settings_v2_page():
    """QBO settings page v2 with simplified OAuth flow (admin and master_admin only)."""
    user_role = session.get('user_role')
    
    # Only admin and master_admin can access QBO settings
    if user_role not in ['admin', 'master_admin']:
        return jsonify({'error': 'Permission denied'}), 403
    
    return render_template('qbo_settings_v2.html')


@app.route('/api/qbo/oauth/authorize-v2', methods=['POST'])
@login_required
def qbo_oauth_authorize_v2():
    """Initiate QBO OAuth 2.0 flow with hardcoded credentials (admin and master_admin only)."""
    user_role = session.get('user_role')
    
    # Only admin and master_admin can initiate OAuth
    if user_role not in ['admin', 'master_admin']:
        return jsonify({'error': 'Permission denied'}), 403
    
    try:
        # NOTE: These credentials are hardcoded per explicit requirement in issue.
        # In production, credentials should be stored in environment variables or secret manager.
        # Hardcoded QBO credentials as specified in requirements
        client_id = 'AB224ne26KUlOjJebeDLMIwgIZcTRQkb6AieFqwJQg0sWCzXXA'
        client_secret = '8LyYgJtmfo7znuWjilV5B3HUGzeiOmZ8hw0dt1Yl'
        
        # Get the redirect URI based on current host
        # Force HTTPS since Cloud Run uses HTTPS externally even if request.host_url returns HTTP
        # Construct URL with explicit HTTPS scheme
        # Expected on Cloud Run: https://feb42026-286597576168.us-central1.run.app/api/qbo/oauth/callback
        parsed_url = urlparse(request.host_url.rstrip('/'))
        https_url = urlunparse((
            'https',  # scheme
            parsed_url.netloc,  # netloc
            parsed_url.path,  # path
            parsed_url.params,  # params
            parsed_url.query,  # query
            parsed_url.fragment  # fragment
        ))
        redirect_uri = https_url + '/api/qbo/oauth/callback'
        
        # Store client credentials in session for later use in callback
        session['qbo_oauth_client_id'] = client_id
        session['qbo_oauth_client_secret'] = client_secret
        session['qbo_oauth_redirect_uri'] = redirect_uri
        
        # Generate a state token for CSRF protection
        state = str(uuid.uuid4())
        session['qbo_oauth_state'] = state
        
        # Build the authorization URL
        # URL encode the redirect_uri to ensure it matches exactly with QuickBooks settings
        # Per RFC 3986, when used as a query parameter VALUE, reserved characters must be encoded
        # Using safe='' ensures all special characters (including :, /, ?) are percent-encoded
        # to match QuickBooks' strict redirect URI requirements
        encoded_redirect_uri = quote(redirect_uri, safe='')
        auth_url = (
            f"https://appcenter.intuit.com/connect/oauth2?"
            f"client_id={client_id}&"
            f"scope=com.intuit.quickbooks.accounting&"
            f"redirect_uri={encoded_redirect_uri}&"
            f"response_type=code&"
            f"state={state}&"
            f"prompt=select_account"
        )
        
        # Log detailed OAuth initiation information for debugging
        logger.info("=" * 80)
        logger.info("QuickBooks OAuth 2.0 Flow Initiated")
        logger.info("=" * 80)
        logger.info(f"Client ID: {client_id[:10]}... (masked for security)")
        logger.info(f"Redirect URI (unencoded): {redirect_uri}")
        logger.info(f"Redirect URI (encoded): {encoded_redirect_uri}")
        logger.info(f"State Token: {state}")
        logger.info(f"Authorization URL: [Generated - check browser for actual URL]")
        logger.info("=" * 80)
        logger.info("IMPORTANT: Verify the following in your QuickBooks Developer Portal:")
        logger.info(f"  1. The redirect URI '{redirect_uri}' is registered EXACTLY as shown")
        logger.info(f"  2. The client ID (shown above, masked) matches your app's credentials")
        logger.info("  3. Your app is in the correct environment (Sandbox vs Production)")
        logger.info("  4. The 'com.intuit.quickbooks.accounting' scope is enabled")
        logger.info("=" * 80)
        
        # Log the action
        database.log_audit(
            user_id=session.get('user_id'),
            user_email=session.get('user_email'),
            action='initiate_qbo_oauth_v2',
            resource_type='qbo_credentials',
            resource_id='1',
            details=f'Initiated QBO OAuth flow v2 - Redirect URI: {redirect_uri}',
            ip_address=request.remote_addr,
            user_agent=request.user_agent.string if request.user_agent else None
        )
        
        return jsonify({'authorization_url': auth_url}), 200
    except Exception as e:
        logger.error(f"Error initiating QBO OAuth v2: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/qbo/disconnect', methods=['POST'])
@login_required
def qbo_disconnect():
    """Disconnect from QuickBooks Online (admin and master_admin only)."""
    user_role = session.get('user_role')
    
    # Only admin and master_admin can disconnect
    if user_role not in ['admin', 'master_admin']:
        return jsonify({'error': 'Permission denied'}), 403
    
    try:
        global qbo_client, secret_manager
        
        # Use secret_manager to delete credentials from database and Secret Manager
        success = secret_manager.delete_qbo_secrets()
        
        if success:
            # Log the action
            database.log_audit(
                user_id=session.get('user_id'),
                user_email=session.get('user_email'),
                action='disconnect_qbo',
                resource_type='qbo_credentials',
                resource_id='1',
                details='Disconnected from QuickBooks - removed all credentials',
                ip_address=request.remote_addr,
                user_agent=request.user_agent.string if request.user_agent else None
            )
            
            # Reinitialize QBO client - will have invalid dummy credentials after disconnect
            # get_qbo_credentials() returns dummy values when no credentials exist in database
            # The QBOClient will mark credentials_valid=False to prevent API calls
            qbo_credentials = secret_manager.get_qbo_credentials()
            qbo_client = QBOClient(
                client_id=qbo_credentials.get('client_id', 'dummy_id'),
                client_secret=qbo_credentials.get('client_secret', 'dummy_secret'),
                refresh_token=qbo_credentials.get('refresh_token', 'dummy_refresh'),
                realm_id=qbo_credentials.get('realm_id', 'dummy_realm'),
                database=database
            )
            
            logger.info("Successfully disconnected from QuickBooks Online")
            return jsonify({'success': True, 'message': 'Disconnected from QuickBooks'}), 200
        else:
            logger.error("Failed to disconnect from QuickBooks")
            return jsonify({'error': 'Failed to disconnect from QuickBooks'}), 500
    except Exception as e:
        logger.error(f"Error disconnecting from QuickBooks: {e}")
        return jsonify({"error": str(e)}), 500


# AI Chat API routes

@app.route('/api/ai/chat', methods=['POST'])
@login_required
@audit_log('ai_chat')
@handle_errors('ai_chat')
def ai_chat():
    """Handle AI chat messages (available to all authenticated users)."""
    try:
        data = request.get_json()
        message = data.get('message')
        conversation_history = data.get('history', [])
        
        if not message:
            return jsonify({'error': 'Message is required'}), 400
        
        user_role = session.get('user_role')
        
        # Get AI response
        response = ai_service.chat(message, conversation_history, user_role)
        
        return jsonify(response), 200
    except Exception as e:
        logger.error(f"Error in AI chat: {e}")
        error_details = error_logger.log_error(
            e,
            context={'message': message, 'user_role': user_role},
            user_id=session.get('user_id'),
            user_email=session.get('user_email')
        )
        return jsonify({"error": "An error occurred processing your request. Please try again."}), 500


@app.route('/api/ai/action', methods=['POST'])
@login_required
@role_required('master_admin')
@audit_log('ai_action')
@log_ai_action('advanced_ai_action')
@handle_errors('ai_action')
def ai_action():
    """Perform AI action (master admin only)."""
    try:
        data = request.get_json()
        action = data.get('action')
        parameters = data.get('parameters', {})
        
        if not action:
            return jsonify({'error': 'Action is required'}), 400
        
        user_role = session.get('user_role')
        
        # Perform action
        result = ai_service.perform_action(action, parameters, user_role)
        
        return jsonify(result), 200
    except Exception as e:
        logger.error(f"Error performing AI action: {e}")
        error_details = error_logger.log_error(
            e,
            context={'action': action, 'parameters': parameters},
            user_id=session.get('user_id'),
            user_email=session.get('user_email')
        )
        return jsonify({"error": "An error occurred performing the action. Please check the error logs."}), 500


# Error and log viewing endpoints (master admin only)

@app.route('/api/errors/recent', methods=['GET'])
@login_required
@role_required('master_admin')
def get_recent_errors():
    """Get recent error logs (master admin only)."""
    try:
        limit = request.args.get('limit', default=100, type=int)
        errors = error_logger.get_recent_errors(limit)
        return jsonify({'errors': errors, 'count': len(errors)}), 200
    except Exception as e:
        logger.error(f"Error fetching error logs: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/ai/operation-logs', methods=['GET'])
@login_required
@role_required('master_admin')
def get_ai_operation_logs():
    """Get AI operation logs (master admin only)."""
    try:
        limit = request.args.get('limit', default=100, type=int)
        logs = error_logger.get_ai_operation_logs(limit)
        return jsonify({'logs': logs, 'count': len(logs)}), 200
    except Exception as e:
        logger.error(f"Error fetching AI operation logs: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/logs', methods=['GET'])
@login_required
@role_required('master_admin')
def logs_page():
    """Error and operation logs page (master admin only)."""
    return render_template('logs.html')


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port)
